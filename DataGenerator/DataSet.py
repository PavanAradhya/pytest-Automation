import openpyxl
li = []
li1 = []

def datagen():
    wk = openpyxl.load_workbook(r"C:/Users/PA061981/Desktop/Multidata.xlsx")
    ws = wk['Sheet1']
    r = ws.max_row
    print(r)
    for i in range(1, r+1):
        li1 = []
        un = ws.cell(i,1)
        us = ws.cell(i,2)
        li1.insert(0, un.value)
        li1.insert(1, us.value)
        li.insert(i-1, li1)
    print(li)
    return li




